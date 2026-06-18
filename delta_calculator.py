import pandas as pd
import numpy as np
from dataclasses import dataclass
from typing import List, Dict, Optional
from openpyxl import load_workbook
import json


@dataclass
class FinancialModel:
    """Структура финансовой модели проекта"""
    project_name: str
    investment_amount: float  # Сумма инвестиций
    project_horizon: int  # Горизонт планирования (лет)
    discount_rate: float  # Ставка дисконтирования (WACC)
    
    # Cash flows по годам
    revenue: List[float]  # Выручка
    opex: List[float]  # Операционные расходы
    capex: List[float]  # Капитальные затраты
    depreciation: List[float]  # Амортизация
    
    # Налоги
    profit_tax_rate: float  # Ставка налога на прибыль (обычно 20%)
    property_tax_rate: float  # Ставка налога на имущество
    land_tax_rate: float  # Ставка земельного налога
    
    # Финансирование
    debt_ratio: float  # Доля заёмного финансирования
    interest_rate: float  # Процентная ставка по кредиту


@dataclass
class SupportMeasure:
    """Структура меры поддержки"""
    name: str
    measure_type: str  # tax_benefit, subsidy, guarantee, infrastructure, loan
    mechanism: str  # profit_tax, property_tax, land_tax, capex_reduction, etc.
    value: float  # Сумма или процент
    value_type: str  # percentage, fixed_amount, rate_reduction
    duration: int  # Срок действия (лет)
    conditions: Dict  # Условия получения
    source_npa: str  # Источник (НПА)


@dataclass
class EconomicEffect:
    """Результат расчёта экономического эффекта"""
    measure_name: str
    baseline_npv: float
    optimized_npv: float
    delta_npv: float
    baseline_irr: float
    optimized_irr: float
    delta_irr: float
    baseline_payback: float
    optimized_payback: float
    delta_payback: float
    roi: float
    details: Dict


class FinancialModelParser:
    """Парсер финансовой модели из Excel"""
    
    @staticmethod
    def parse_excel(file_path: str) -> FinancialModel:
        """
        Парсит финансовую модель из Excel файла
        """
        wb = load_workbook(file_path, data_only=True)
        
        # Чтение листа с основными показателями
        ws = wb['Основные показатели']  # или другой лист
        
        # Извлечение данных (пример - структура может отличаться)
        model = FinancialModel(
            project_name=ws['B2'].value,
            investment_amount=float(ws['B5'].value),
            project_horizon=int(ws['B6'].value),
            discount_rate=float(ws['B7'].value) / 100,
            
            revenue=[float(ws.cell(row=10, column=i).value) for i in range(2, 12)],
            opex=[float(ws.cell(row=11, column=i).value) for i in range(2, 12)],
            capex=[float(ws.cell(row=12, column=i).value) for i in range(2, 12)],
            depreciation=[float(ws.cell(row=13, column=i).value) for i in range(2, 12)],
            
            profit_tax_rate=0.20,  # Стандартная ставка
            property_tax_rate=0.022,  # 2.2%
            land_tax_rate=0.015,  # 1.5%
            
            debt_ratio=0.7,
            interest_rate=0.12
        )
        
        return model


class SupportMeasureClassifier:
    """Классификатор мер поддержки по типу воздействия"""
    
    @staticmethod
    def classify(measure: SupportMeasure) -> str:
        """
        Определяет, как мера влияет на финансовую модель
        """
        impact_map = {
            'tax_benefit': {
                'profit_tax': 'reduce_profit_tax',
                'property_tax': 'reduce_property_tax',
                'land_tax': 'reduce_land_tax',
            },
            'subsidy': {
                'capex': 'reduce_capex',
                'opex': 'reduce_opex',
            },
            'guarantee': 'reduce_interest_rate',
            'infrastructure': 'reduce_capex',
            'loan': 'reduce_interest_rate',
        }
        
        if measure.measure_type == 'tax_benefit':
            return impact_map['tax_benefit'].get(measure.mechanism, 'unknown')
        elif measure.measure_type == 'subsidy':
            return impact_map['subsidy'].get(measure.mechanism, 'unknown')
        else:
            return impact_map.get(measure.measure_type, 'unknown')


class ImpactCalculator:
    """Калькулятор воздействия мер поддержки на финансовую модель"""
    
    def __init__(self, model: FinancialModel):
        self.model = model
        self.classifier = SupportMeasureClassifier()
    
    def apply_measure(self, measure: SupportMeasure) -> FinancialModel:
        """
        Применяет меру поддержки к финансовой модели
        Возвращает модифицированную модель
        """
        # Создаём копию модели
        modified = FinancialModel(
            project_name=self.model.project_name,
            investment_amount=self.model.investment_amount,
            project_horizon=self.model.project_horizon,
            discount_rate=self.model.discount_rate,
            revenue=self.model.revenue.copy(),
            opex=self.model.opex.copy(),
            capex=self.model.capex.copy(),
            depreciation=self.model.depreciation.copy(),
            profit_tax_rate=self.model.profit_tax_rate,
            property_tax_rate=self.model.property_tax_rate,
            land_tax_rate=self.model.land_tax_rate,
            debt_ratio=self.model.debt_ratio,
            interest_rate=self.model.interest_rate
        )
        
        impact_type = self.classifier.classify(measure)
        
        # Применение воздействия
        if impact_type == 'reduce_profit_tax':
            if measure.value_type == 'rate_reduction':
                modified.profit_tax_rate -= measure.value / 100
            elif measure.value_type == 'percentage':
                modified.profit_tax_rate *= (1 - measure.value / 100)
        
        elif impact_type == 'reduce_property_tax':
            if measure.value_type == 'percentage':
                modified.property_tax_rate *= (1 - measure.value / 100)
        
        elif impact_type == 'reduce_land_tax':
            if measure.value_type == 'percentage':
                modified.land_tax_rate *= (1 - measure.value / 100)
        
        elif impact_type == 'reduce_capex':
            if measure.value_type == 'fixed_amount':
                # Распределяем сумму по годам (обычно в первые годы)
                years_to_apply = min(measure.duration, len(modified.capex))
                annual_reduction = measure.value / years_to_apply
                for i in range(years_to_apply):
                    modified.capex[i] -= annual_reduction
            elif measure.value_type == 'percentage':
                for i in range(min(measure.duration, len(modified.capex))):
                    modified.capex[i] *= (1 - measure.value / 100)
        
        elif impact_type == 'reduce_opex':
            if measure.value_type == 'fixed_amount':
                annual_reduction = measure.value / measure.duration
                for i in range(min(measure.duration, len(modified.opex))):
                    modified.opex[i] -= annual_reduction
            elif measure.value_type == 'percentage':
                for i in range(min(measure.duration, len(modified.opex))):
                    modified.opex[i] *= (1 - measure.value / 100)
        
        elif impact_type == 'reduce_interest_rate':
            if measure.value_type == 'rate_reduction':
                modified.interest_rate -= measure.value / 100
            elif measure.value_type == 'percentage':
                modified.interest_rate *= (1 - measure.value / 100)
        
        return modified


class DCFModel:
    """Модель дисконтированных денежных потоков"""
    
    @staticmethod
    def calculate_cash_flows(model: FinancialModel) -> List[float]:
        """
        Расчёт свободных денежных потоков (FCFF)
        """
        cash_flows = []
        
        for year in range(model.project_horizon):
            # Выручка
            revenue = model.revenue[year] if year < len(model.revenue) else 0
            
            # Операционные расходы
            opex = model.opex[year] if year < len(model.opex) else 0
            
            # Амортизация
            depreciation = model.depreciation[year] if year < len(model.depreciation) else 0
            
            # EBITDA
            ebitda = revenue - opex
            
            # EBIT
            ebit = ebitda - depreciation
            
            # Налог на прибыль
            profit_tax = max(0, ebit * model.profit_tax_rate)
            
            # Налоги на имущество и землю (упрощённо)
            property_tax = model.investment_amount * model.property_tax_rate / model.project_horizon
            land_tax = model.investment_amount * 0.01 * model.land_tax_rate / model.project_horizon
            
            # Чистая прибыль
            net_income = ebit - profit_tax - property_tax - land_tax
            
            # Проценты по кредиту
            debt_amount = model.investment_amount * model.debt_ratio
            interest_expense = debt_amount * model.interest_rate
            
            # Операционный денежный поток
            operating_cf = net_income + depreciation
            
            # Капитальные затраты
            capex = model.capex[year] if year < len(model.capex) else 0
            
            # Свободный денежный поток
            fcf = operating_cf - capex - interest_expense * (1 - model.profit_tax_rate)
            
            # В год 0 добавляем инвестиции
            if year == 0:
                fcf -= model.investment_amount * (1 - model.debt_ratio)
            
            cash_flows.append(fcf)
        
        return cash_flows
    
    @staticmethod
    def calculate_npv(cash_flows: List[float], discount_rate: float) -> float:
        """Расчёт NPV"""
        npv = sum(cf / (1 + discount_rate) ** t for t, cf in enumerate(cash_flows))
        return npv
    
    @staticmethod
    def calculate_irr(cash_flows: List[float]) -> Optional[float]:
        """Расчёт IRR"""
        try:
            from scipy.optimize import brentq
            
            def npv_function(rate):
                return sum(cf / (1 + rate) ** t for t, cf in enumerate(cash_flows))
            
            irr = brentq(npv_function, -0.5, 1.0)
            return irr
        except:
            return None
    
    @staticmethod
    def calculate_payback(cash_flows: List[float]) -> float:
        """Расчёт срока окупаемости"""
        cumulative_cf = 0
        for year, cf in enumerate(cash_flows):
            cumulative_cf += cf
            if cumulative_cf >= 0:
                # Интерполяция для точного срока
                if year > 0:
                    prev_cf = cumulative_cf - cf
                    fraction = -prev_cf / cf if cf != 0 else 0
                    return year - 1 + fraction
                return year
        return float('inf')


class EconomicEffectCalculator:
    """Основной калькулятор экономического эффекта"""
    
    def __init__(self, model: FinancialModel):
        self.model = model
        self.impact_calc = ImpactCalculator(model)
        self.dcf = DCFModel()
    
    def calculate_baseline(self) -> Dict:
        """Расчёт базового сценария (без мер поддержки)"""
        cash_flows = self.dcf.calculate_cash_flows(self.model)
        
        return {
            'cash_flows': cash_flows,
            'npv': self.dcf.calculate_npv(cash_flows, self.model.discount_rate),
            'irr': self.dcf.calculate_irr(cash_flows),
            'payback': self.dcf.calculate_payback(cash_flows),
        }
    
    def calculate_with_measure(self, measure: SupportMeasure) -> EconomicEffect:
        """Расчёт эффекта от одной меры поддержки"""
        # Базовый сценарий
        baseline = self.calculate_baseline()
        
        # Модифицированная модель с мерой
        modified_model = self.impact_calc.apply_measure(measure)
        modified_flows = self.dcf.calculate_cash_flows(modified_model)
        
        optimized_npv = self.dcf.calculate_npv(modified_flows, self.model.discount_rate)
        optimized_irr = self.dcf.calculate_irr(modified_flows)
        optimized_payback = self.dcf.calculate_payback(modified_flows)
        
        # Расчёт дельт
        delta_npv = optimized_npv - baseline['npv']
        delta_irr = (optimized_irr - baseline['irr']) if baseline['irr'] and optimized_irr else None
        delta_payback = baseline['payback'] - optimized_payback
        
        # ROI
        roi = delta_npv / self.model.investment_amount if self.model.investment_amount > 0 else 0
        
        return EconomicEffect(
            measure_name=measure.name,
            baseline_npv=baseline['npv'],
            optimized_npv=optimized_npv,
            delta_npv=delta_npv,
            baseline_irr=baseline['irr'] or 0,
            optimized_irr=optimized_irr or 0,
            delta_irr=delta_irr or 0,
            baseline_payback=baseline['payback'],
            optimized_payback=optimized_payback,
            delta_payback=delta_payback,
            roi=roi,
            details={
                'measure_type': measure.measure_type,
                'mechanism': measure.mechanism,
                'value': measure.value,
                'baseline_cash_flows': baseline['cash_flows'],
                'optimized_cash_flows': modified_flows,
            }
        )
    
    def calculate_cumulative_effect(self, measures: List[SupportMeasure]) -> Dict:
        """
        Расчёт кумулятивного эффекта от всех мер поддержки
        """
        # Базовый сценарий
        baseline = self.calculate_baseline()
        
        # Применяем все меры последовательно
        cumulative_model = self.model
        for measure in measures:
            cumulative_model = self.impact_calc.apply_measure(measure)
        
        cumulative_flows = self.dcf.calculate_cash_flows(cumulative_model)
        
        optimized_npv = self.dcf.calculate_npv(cumulative_flows, self.model.discount_rate)
        optimized_irr = self.dcf.calculate_irr(cumulative_flows)
        optimized_payback = self.dcf.calculate_payback(cumulative_flows)
        
        delta_npv = optimized_npv - baseline['npv']
        delta_irr = (optimized_irr - baseline['irr']) if baseline['irr'] and optimized_irr else 0
        delta_payback = baseline['payback'] - optimized_payback
        roi = delta_npv / self.model.investment_amount
        
        # Индивидуальные эффекты
        individual_effects = [self.calculate_with_measure(m) for m in measures]
        
        return {
            'baseline': baseline,
            'cumulative': {
                'npv': optimized_npv,
                'irr': optimized_irr,
                'payback': optimized_payback,
                'cash_flows': cumulative_flows,
            },
            'delta': {
                'npv': delta_npv,
                'irr': delta_irr,
                'payback': delta_payback,
                'roi': roi,
            },
            'individual_effects': individual_effects,
            'total_measures': len(measures),
        }


class ReportGenerator:
    """Генератор отчёта по результатам расчёта"""
    
    @staticmethod
    def generate_summary(results: Dict) -> str:
        """Генерация текстового summary для руководства"""
        baseline = results['baseline']
        cumulative = results['cumulative']
        delta = results['delta']
        
        summary = f"""
=== АНАЛИЗ ЭКОНОМИЧЕСКОГО ЭФФЕКТА ОТ МЕР ПОДДЕРЖКИ ===

Проект: {results.get('project_name', 'Не указан')}
Количество подобранных мер: {results['total_measures']}

БАЗОВЫЙ СЦЕНАРИЙ (без мер поддержки):
• NPV: {baseline['npv']:,.0f} руб.
• IRR: {baseline['irr']*100:.1f}%
• Срок окупаемости: {baseline['payback']:.1f} лет

ОПТИМИЗИРОВАННЫЙ СЦЕНАРИЙ (с мерами поддержки):
• NPV: {cumulative['npv']:,.0f} руб.
• IRR: {cumulative['irr']*100:.1f}%
• Срок окупаемости: {cumulative['payback']:.1f} лет

ЭКОНОМИЧЕСКИЙ ЭФФЕКТ:
• Прирост NPV: {delta['npv']:,.0f} руб. ({delta['npv']/baseline['npv']*100:.1f}%)
• Прирост IRR: {delta['irr']*100:.1f} п.п.
• Сокращение срока окупаемости: {delta['payback']:.1f} лет
• ROI от мер поддержки: {delta['roi']*100:.1f}%

ТОП-3 МЕРЫ ПО ВЛИЯНИЮ НА NPV:
"""
        
        # Сортировка индивидуальных эффектов
        sorted_effects = sorted(
            results['individual_effects'],
            key=lambda x: x.delta_npv,
            reverse=True
        )
        
        for i, effect in enumerate(sorted_effects[:3], 1):
            summary += f"\n{i}. {effect.measure_name}"
            summary += f"\n   • ΔNPV: {effect.delta_npv:,.0f} руб."
            summary += f"\n   • ΔIRR: {effect.delta_irr*100:.1f} п.п."
            summary += f"\n   • ΔPayback: {effect.delta_payback:.1f} лет"
        
        return summary
    
    @staticmethod
    def generate_detailed_table(results: Dict) -> pd.DataFrame:
        """Генерация детальной таблицы эффектов"""
        rows = []
        for effect in results['individual_effects']:
            rows.append({
                'Мера поддержки': effect.measure_name,
                'Тип': effect.details['measure_type'],
                'Механизм': effect.details['mechanism'],
                'Базовый NPV': f"{effect.baseline_npv:,.0f}",
                'NPV с мерой': f"{effect.optimized_npv:,.0f}",
                'ΔNPV': f"{effect.delta_npv:,.0f}",
                'ΔIRR (%)': f"{effect.delta_irr*100:.2f}",
                'ΔPayback (лет)': f"{effect.delta_payback:.1f}",
                'ROI (%)': f"{effect.roi*100:.1f}",
            })
        
        return pd.DataFrame(rows)


# Пример использования
def main():
    """Пример использования инструмента"""
    
    # 1. Загрузка финансовой модели
    model = FinancialModelParser.parse_excel('financial_model.xlsx')
    
    # 2. Загрузка мер поддержки (из JSON или API)
    with open('support_measures.json', 'r', encoding='utf-8') as f:
        measures_data = json.load(f)
    
    measures = [SupportMeasure(**m) for m in measures_data]
    
    # 3. Расчёт эффектов
    calculator = EconomicEffectCalculator(model)
    results = calculator.calculate_cumulative_effect(measures)
    
    # 4. Генерация отчёта
    summary = ReportGenerator.generate_summary(results)
    print(summary)
    
    # 5. Экспорт детальной таблицы
    df = ReportGenerator.generate_detailed_table(results)
    df.to_excel('economic_effects_report.xlsx', index=False)
    
    return results


if __name__ == '__main__':
    main()
